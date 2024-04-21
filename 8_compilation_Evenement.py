import os
import openpyxl
import locale
from datetime import datetime, timedelta

# Paramétrer la locale pour français
locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8' if os.name == 'nt' else 'fr_FR.utf8')

# Chemins des fichiers
source_path = "C:/Users/x/Desktop/NeuroPal_Global/Base_de_donnees/3_Extraction_donnees/2_Extraction_evenement/Evenements.xlsx"
destination_path = "C:/Users/x/Desktop/NeuroPal_Global/Base_de_donnees/4_Donnees_interface/0_calendar/dates_and_days_with_week_number_2024.xlsx"

# Charger les workbooks
source_wb = openpyxl.load_workbook(source_path)
destination_wb = openpyxl.load_workbook(destination_path)

# Charger les feuilles
source_ws = source_wb.active
destination_ws = destination_wb.active

# Fonction pour interpréter des termes relatifs comme 'demain'
def interpret_relative_date(term, base_date=datetime.today()):
    term = term.lower()
    if 'demain' in term:
        return base_date + timedelta(days=1)
    if 'hier' in term:
        return base_date - timedelta(days=1)
    # Ajouter d'autres termes relatifs et leur logique ici
    return None

# Dictionnaire pour associer les dates et les descriptions
date_description_dict = {}

# Lire les données du fichier source et les stocker dans le dictionnaire
for row in source_ws.iter_rows(min_row=2, values_only=True):
    date_text = str(row[0])
    try:
        date = datetime.strptime(date_text, "%d %B").replace(year=2024)
    except ValueError:
        date = interpret_relative_date(date_text)
        if date is None:
            print(f"La date '{date_text}' n'est pas reconnue ou n'est pas gérée par le script.")
            continue

    description = row[1]
    if date not in date_description_dict:
        date_description_dict[date] = [description]
    else:
        date_description_dict[date].append(description)

# Parcourir le fichier de destination et ajouter les descriptions correspondantes
for row_index, row_data in enumerate(destination_ws.iter_rows(min_row=2, values_only=True), start=2):
    cell_date = row_data[1]  # La date est dans la deuxième colonne du tuple
    if cell_date in date_description_dict:
        descriptions = date_description_dict[cell_date]
        initial_col = 16  # Colonne initiale pour les descriptions
        for description in descriptions:
            # Trouver la première cellule vide ou la suivante disponible
            col_offset = 0
            while destination_ws.cell(row=row_index, column=initial_col + col_offset).value is not None:
                col_offset += 1
            cell_to_write = destination_ws.cell(row=row_index, column=initial_col + col_offset)
            cell_to_write.value = description

# Sauvegarder le fichier de destination
destination_wb.save(destination_path)
