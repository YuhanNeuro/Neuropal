import openpyxl
from openpyxl.utils import get_column_letter

# Chemins vers les fichiers source et destination
src_path = r'C:\Users\x\Desktop\NeuroPal_Global\Base_de_donnees\2_Historiques\history_4semaines\4semaine.xlsx'
dest_path = r'C:\Users\x\Desktop\NeuroPal_Global\Base_de_donnees\4_Donnees_interface\0_calendar\Diapo_reposit_vierge.xlsx'

# Charger les workbooks
src_book = openpyxl.load_workbook(src_path)
dest_book = openpyxl.load_workbook(dest_path)

# Charger les feuilles actives
src_sheet = src_book.active
dest_sheet = dest_book.active

# Plage à copier de la source
start_col = 'A'
start_row = 1
end_col = 'Z'
end_row = 32

# Cellule de début dans le fichier de destination
dest_start_col = 'N'
dest_start_row = 26

# Copie des cellules
for row in range(start_row, end_row + 1):
    for col in range(ord(start_col), ord(end_col) + 1):
        src_cell = f"{chr(col)}{row}"
        dest_col_index = ord(dest_start_col) - ord('A') + (col - ord(start_col)) + 1
        dest_cell = f"{get_column_letter(dest_col_index)}{dest_start_row + (row - start_row)}"
        dest_sheet[dest_cell].value = src_sheet[src_cell].value

# Sauvegarder le fichier de destination
dest_book.save(dest_path)

print("Les données ont été copiées avec succès.")
